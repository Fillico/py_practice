# -*- coding:utf-8 -*-
#@author:Fillico
#@date:2019/4/15 9:20

from openpyxl import load_workbook
from python_api_3.common.http_request import HttpSessionRequest

class Cases:
    def __init__(self):
        self.case_id = None
        self.title = None
        self.url = None
        self.data = None
        self.method = None
        self.expected = None
        self.actual = None
        self.result = None
        self.sql = None

class DoExcel:
    def __init__(self,file_name,sheet_name):
        '''
        #定义一个类，可以读写Excel
        :param file_name:需要操作的Excel文件名
        :param sheet_name: 需要操作的表单
        '''
        try:
            self.filename=file_name
            self.sheet_name=sheet_name
            self.workbook = load_workbook(file_name)
            self.sheet=self.workbook[sheet_name]
        except Exception as e:
            print("文件不存在")

    def get_cases(self):
        max_row = self.sheet.max_row
        cases=[]
        try:
            for row in range(2,max_row+1):
                case={}
                case = Cases()
                case.case_id = self.sheet.cell(row,column=1).value
                case.title = self.sheet.cell(row,column=2).value
                case.url = self.sheet.cell(row,column=3).value
                case.data = self.sheet.cell(row,column=4).value
                case.method = self.sheet.cell(row,column=5).value
                case.expected = self.sheet.cell(row,column=6).value
                case.sql = self.sheet.cell(row,column=9).value
                cases.append(case)
            self.workbook.close()
        except Exception as e:
            raise e
        return cases

    def write_result(self,row_num,actual,result):
        # sheet=self.workbook.sheet[self.sheet_name]
        self.sheet.cell(row_num,7).value=actual
        self.sheet.cell(row_num,8).value=result
        # wb = load_workbook(filename=self.filename)
        # sheet = wb[self.sheetname]
        # sheet.cell(row_num,column_num).value=content
        self.workbook.save(filename=self.filename)
        self.workbook.close()

if __name__ == '__main__':
    do_excel=DoExcel(r'python_api_3\data\future_cases.xlsx','register')
    cases = do_excel.get_cases()
    # for case in cases:
    #     # print(case.__dict__)
    #     resp_case = HttpSessionRequest().session_request(case.method,case.url,eval(case.data))
    #     actual=resp_case.text
    #     # print(actual)
    #     if case.expected == actual:
    #         do_excel.write_result(case.case_id+1,actual,"PASS")
    #     else:
    #         do_excel.write_result(case.case_id+1,actual,"FAIL")
    print(cases)






