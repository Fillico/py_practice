# -*- coding:utf-8 -*-
#@author:Fillico
#@date:2019/3/25 15:51

from openpyxl import load_workbook
from week6.lesson_0328.learn_config import MyConfig

class ReadWriteExcel:
    def __init__(self,filename,sheetname):
        self.filename=filename
        self.sheetname=sheetname
    def read_excel(self):
        wb=load_workbook(self.filename)
        sheet=wb[self.sheetname]
        #从配置文件读取数据，决定获取哪些行的数据
        line = MyConfig('les_0319.conf').get_strValue('LINENO', 'line')
        data_all=[]
        for i in range(1,sheet.max_row+1):
            data_line=[]
            for j in range(1,sheet.max_column+1):
                if sheet.cell(i,j).value:
                    data_line.append(sheet.cell(i,j).value)
            data_all.append(data_line)
        final_data=[] #最后返回的数据
        if line=='all':#读取所有行的数据
            final_data=data_all
        else:#读取指定列表里面指定行的数据
            for i in eval(line):#遍历列表里面行数的数字，  eval(line)把字符串的line转换成列表的line
                final_data.append(data_all[i-1])
        wb.close()
        return final_data

    def write_excel(self,row,column,new_value):
        wb=load_workbook(self.filename)
        sheet=wb[self.sheetname]
        sheet.cell(row,column).value=new_value
        wb.save(self.filename)
        wb.close()
if __name__ == '__main__':

    operate_excel=ReadWriteExcel('login_cases.xlsx','Sheet1')
    # operate_excel.write_excel(2,2,'Fillico')
    res=operate_excel.read_excel()
    print(res)



